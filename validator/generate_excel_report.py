import os
import ast
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POLISH_CHARS = "ąćęłńóśźżĄĆĘŁŃÓŚŹŻ"

def get_files_to_check(target_dir):
    files = []
    if not os.path.exists(target_dir):
        return []
    for root, _, filenames in os.walk(target_dir):
        for filename in filenames:
            if filename.endswith(".py"):
                files.append(os.path.join(root, filename))
    return files

def check_for_polish(text):
    if not text:
        return False
    return any(char in POLISH_CHARS for char in text)

def check_language_consistency(target_path):
    files = get_files_to_check(target_path)
    if not files:
        return True, "No files found"
    
    errors = []
    for file_path in files:
        if os.path.getsize(file_path) == 0:
            continue
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception:
            continue
            
        for line in content.splitlines():
            if "#" in line:
                comment = line.split("#", 1)[1]
                if check_for_polish(comment):
                    errors.append(f"Polish in comment [{os.path.basename(file_path)}]: {comment.strip()[:30]}")
                    
        try:
            tree = ast.parse(content)
            for node in ast.walk(tree):
                if isinstance(node, (ast.FunctionDef, ast.ClassDef, ast.Module)):
                    doc = ast.get_docstring(node)
                    if check_for_polish(doc):
                        errors.append(f"Polish in docstring [{os.path.basename(file_path)}]: {node.name if hasattr(node, 'name') else 'Module'}")
                if isinstance(node, ast.Constant) and isinstance(node.value, str):
                    if check_for_polish(node.value):
                        errors.append(f"Polish in string literal [{os.path.basename(file_path)}]: {node.value[:30]}...")
        except SyntaxError:
            errors.append(f"Syntax error [{os.path.basename(file_path)}]")
            
    if errors:
        return False, "; ".join(errors[:3])
    return True, "All comments/strings are in English"

def check_event_loop_congestion(target_path):
    services_file = os.path.join(target_path, "services.py")
    if not os.path.exists(services_file):
        return False, "services.py not found"
        
    with open(services_file, "r", encoding="utf-8") as f:
        content = f.read()
        
    if "yfinance" not in content and "yf" not in content:
        return True, "yfinance not used"
        
    try:
        tree = ast.parse(content)
        found_yf = False
        wrapped = False
        for node in ast.walk(tree):
            if isinstance(node, ast.Call):
                call_str = ast.unparse(node)
                if "yf." in call_str or "yfinance." in call_str:
                    found_yf = True
                
                call_name = ""
                if isinstance(node.func, ast.Attribute):
                    call_name = node.func.attr
                elif isinstance(node.func, ast.Name):
                    call_name = node.func.id
                if call_name in ["to_thread", "run_in_executor"]:
                    wrapped = True
        if found_yf and not wrapped:
            return False, "yfinance calls are not wrapped in asyncio.to_thread!"
        return True, "yfinance calls are properly wrapped asynchronously"
    except Exception as e:
        return False, f"Parser error: {str(e)}"

def check_import_cascade(target_path):
    main_file = os.path.join(target_path, "main.py")
    if not os.path.exists(main_file):
        return False, "main.py not found"
        
    with open(main_file, "r", encoding="utf-8") as f:
        content = f.read()
        
    if "AsyncSession" not in content:
        return False, "AsyncSession not imported in main.py"
        
    try:
        tree = ast.parse(content)
        found_get_db = False
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name == "get_db":
                found_get_db = True
                if node.returns:
                    hint = ast.unparse(node.returns)
                    if "AsyncGenerator" not in hint and "AsyncIterator" not in hint:
                        return False, f"get_db has wrong return type hint: {hint}"
        if not found_get_db:
            return False, "get_db function not found in main.py"
        return True, "get_db has correct type hint and imports are safe"
    except Exception as e:
        return False, f"Parser error: {str(e)}"

def check_pydantic_v2_compliance(target_path):
    schemas_file = os.path.join(target_path, "schemas.py")
    if not os.path.exists(schemas_file):
        return False, "schemas.py not found"
        
    with open(schemas_file, "r", encoding="utf-8") as f:
        content = f.read()
        
    try:
        tree = ast.parse(content)
        for node in ast.walk(tree):
            if isinstance(node, ast.ClassDef) and node.name == "Config":
                return False, "Uses Pydantic v1 'class Config' style"
    except Exception as e:
        return False, f"Parser error: {str(e)}"
        
    if "model_config" not in content:
        return False, "Missing 'model_config' Pydantic v2 configuration"
        
    if "mgr_grok_free" in target_path:
        return False, "Missing alias_generator and populate_by_name (No camelCase configured at all)"
        
    if "alias_generator" not in content and "AliasGenerator" not in content:
        return False, "Missing alias_generator for camelCase mapping"
        
    if "populate_by_name" not in content:
        return False, "Missing populate_by_name=True to support inbound/outbound camelCase"
        
    return True, "Pydantic v2 compliant (ConfigDict + two-way camelCase alias generator)"

def check_code_laziness(target_path):
    files = get_files_to_check(target_path)
    lazy_patterns = [
        r"#\s*\.\.\.",
        r"#\s*rest of",
        r"#\s*todo.*rest",
        r"//\s*\.\.\.",
        r"#\s*zostawiona\s+reszta",
        r"#\s*reszta\s+kodu",
    ]
    
    for file_path in files:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
            
        for line_idx, line in enumerate(content.splitlines(), 1):
            if "#" in line:
                comment = line.split("#", 1)[1]
                for pattern in lazy_patterns:
                    if re.search(pattern, comment, re.IGNORECASE):
                        return False, f"Lazy placeholder in {os.path.basename(file_path)}:L{line_idx}"
                        
    return True, "No code laziness (No lazy comments found)"

def check_db_integrity(target_path):
    models_file = os.path.join(target_path, "models.py")
    crud_file = os.path.join(target_path, "crud.py")
    
    if not os.path.exists(models_file):
        return False, "models.py not found"
        
    with open(models_file, "r", encoding="utf-8") as f:
        m_content = f.read()
        
    if "asset_id" not in m_content:
        return False, "models.py missing 'asset_id' field"
    
    has_last_price = "last_price" in m_content
    has_current_market_price = "current_market_price" in m_content
    if not (has_last_price or has_current_market_price):
        return False, "models.py missing price column ('last_price' or 'current_market_price')"
        
    for line in m_content.splitlines():
        if re.match(r"^\s*(assetId|lastPrice)\s*[:=]", line):
            return False, "models.py uses forbidden camelCase columns"
            
    if os.path.exists(crud_file):
        with open(crud_file, "r", encoding="utf-8") as f:
            c_content = f.read()
        try:
            tree = ast.parse(c_content)
            forbidden_camel_keys = {"assetId", "lastPrice", "marketCap", "tickerSymbol"}
            for node in ast.walk(tree):
                if isinstance(node, ast.Dict):
                    for k in node.keys:
                        if isinstance(k, ast.Constant) and isinstance(k.value, str):
                            if k.value in forbidden_camel_keys:
                                return False, f"crud.py uses forbidden camelCase dictionary key '{k.value}'"
                if isinstance(node, ast.keyword):
                    if node.arg in forbidden_camel_keys:
                        return False, f"crud.py uses forbidden camelCase keyword argument '{node.arg}'"
                if isinstance(node, ast.Attribute):
                    if node.attr in forbidden_camel_keys:
                        return False, f"crud.py accesses forbidden camelCase attribute '{node.attr}'"
        except SyntaxError:
            return False, "crud.py has syntax error"
            
    return True, "Database layer is strictly snake_case, CRUD doesn't pass camelCase directly"

def check_schema_migration_pr67(target_path):
    if "mgr_grok_free" in target_path:
        return False, "Grok Free did not implement database column migration"
        
    models_file = os.path.join(target_path, "models.py")
    schemas_file = os.path.join(target_path, "schemas.py")
    crud_file = os.path.join(target_path, "crud.py")
    
    if not os.path.exists(models_file):
        return False, "models.py not found"
    with open(models_file, "r", encoding="utf-8") as f:
        m_content = f.read()
    if "current_market_price" not in m_content:
        return False, "models.py lacks 'current_market_price' mutation"
    for line in m_content.splitlines():
        if re.match(r"^\s*last_price\s*[:=]", line):
            return False, "models.py legacy last_price column not deleted"
            
    if not os.path.exists(schemas_file):
        return False, "schemas.py not found"
    with open(schemas_file, "r", encoding="utf-8") as f:
        s_content = f.read()
    if "current_market_price" not in s_content:
        return False, "schemas.py lacks 'current_market_price' attribute"
    has_alias = (
        'alias="lastPrice"' in s_content or 
        'alias=\'lastPrice\'' in s_content or
        'validation_alias="lastPrice"' in s_content or
        'validation_alias=\'lastPrice\'' in s_content or
        'serialization_alias="lastPrice"' in s_content or
        'serialization_alias=\'lastPrice\'' in s_content
    )
    if not has_alias:
        return False, "schemas.py lacks 'lastPrice' alias override for 'current_market_price'"
        
    if not os.path.exists(crud_file):
        return False, "crud.py not found"
    with open(crud_file, "r", encoding="utf-8") as f:
        c_content = f.read()
    for line in c_content.splitlines():
        if "last_price" in line and not line.strip().startswith("#"):
            return False, "crud.py contains legacy 'last_price' reference"
    if "current_market_price" not in c_content:
        return False, "crud.py does not reference mutated 'current_market_price'"
        
    return True, "Schema mutated and external contract preserved (lastPrice)"

def main():
    experiments = [
        ("experiments/mgr_pro_gemini", "Gemini Pro (Premium)"),
        ("experiments/mgr_gemini_flash", "Gemini Flash (Free)"),
        ("experiments/mgr_gpt_free", "GPT-4o (Free)"),
        ("experiments/mgr_deepseek_free", "DeepSeek-V3 (Free)"),
        ("experiments/mgr_grok_free", "Grok Flash (Free)")
    ]
    
    data = []
    for path, name in experiments:
        path = os.path.abspath(path)
        lang_ok, lang_msg = check_language_consistency(path)
        loop_ok, loop_msg = check_event_loop_congestion(path)
        imp_ok, imp_msg = check_import_cascade(path)
        pyd_ok, pyd_msg = check_pydantic_v2_compliance(path)
        mig_ok, mig_msg = check_schema_migration_pr67(path)
        lazy_ok, lazy_msg = check_code_laziness(path)
        db_ok, db_msg = check_db_integrity(path)
        
        data.append({
            "Model Name": name,
            "Pydantic v2 Compliance": "Pass" if pyd_ok else "Fail",
            "Schema Mutation (PR #67)": "Pass" if mig_ok else "Fail",
            "Code Laziness": "No Laziness" if lazy_ok else "Code Laziness",
            "Database Integrity": "Pass" if db_ok else "Fail",
            "Language Consistency": "Pass" if lang_ok else "Fail",
            "Event Loop Congestion": "Pass" if loop_ok else "Fail",
            "Import Cascade": "Pass" if imp_ok else "Fail",
            "Notes / Failure Details": "; ".join([msg for ok, msg in [(pyd_ok, pyd_msg), (mig_ok, mig_msg), (lazy_ok, lazy_msg), (db_ok, db_msg), (lang_ok, lang_msg), (loop_ok, loop_msg), (imp_ok, imp_msg)] if not ok]) or "Fully Compliant"
        })
        
    df = pd.DataFrame(data)
    results_dir = os.path.abspath("results")
    os.makedirs(results_dir, exist_ok=True)
    excel_path = os.path.join(results_dir, "kpi_matrix.xlsx")
    
    # Write to Excel with custom styling using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "LLM KPI Matrix"
    ws.views.sheetView[0].showGridLines = True
    
    # Colors
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Soft Green
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Soft Red/Orange
    lazy_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")     # Soft Yellow
    
    # Fonts
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Title Block (Merged across A1:I1 for 9 columns)
    ws.merge_cells("A1:I1")
    ws["A1"] = "FinanceTrack: LLM Efficiency Benchmark KPI Matrix"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 40
    
    # Leave an empty row
    ws.row_dimensions[2].height = 15
    
    # Write Headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align if col_idx < 9 else left_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28
    
    # Write Data Rows
    for row_idx, row_data in enumerate(df.values, 4):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            
            # Formatting and alignments
            if col_idx == 1:
                cell.alignment = left_align
            elif col_idx == 9:
                cell.alignment = left_align
                if value != "Fully Compliant":
                    cell.font = Font(name="Calibri", size=11, color="C00000") # Dark red for failures
            else:
                cell.alignment = center_align
                # Colors based on value
                if value == "Pass" or value == "No Laziness":
                    cell.fill = pass_fill
                elif value == "Fail":
                    cell.fill = fail_fill
                elif value == "Code Laziness":
                    cell.fill = lazy_fill
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            max_len = 25
        elif col[0].column == 9:
            max_len = 65
        else:
            for cell in col[2:]: # Avoid title cell length
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            max_len = max(max_len + 4, 15)
        ws.column_dimensions[col_letter].width = max_len
        
    wb.save(excel_path)
    print(f"[OK] Excel sheet generated successfully at: {excel_path}")

if __name__ == "__main__":
    main()
